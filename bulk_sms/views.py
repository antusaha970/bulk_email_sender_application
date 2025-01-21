from django.shortcuts import render

from rest_framework.views import APIView
from .serializers import *
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction
from config.tasks import send_sms_task
from twilio.rest import Client
from rest_framework.permissions import IsAuthenticated

from openpyxl import load_workbook
from io import BytesIO
from django.core.exceptions import ValidationError


class SmsConfigurationView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        print("welcome")
        username = request.data.get("username")
        account_sid = request.data.get("account_sid")
        auth_token = request.data.get("auth_token")
        sender_number = request.data.get("sender_number")
        user = request.user
        print(username, account_sid, auth_token, sender_number, user)
        if not all([username, account_sid, auth_token, sender_number]):
            return Response(
                {"errors": "username, account_sid, auth_token, sender_number,are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        print("previous")
        serializer = SmsConfigurationSerializer(data={
            "username": username,
            "account_sid": account_sid,
            "auth_token": auth_token,
            "sender_number": sender_number,
            "user": user.id
        })
        print("after")
        if serializer.is_valid():
            sms_config = serializer.save()
            return Response({
                "status": "success",
                "details": "Successfully added Twilio Configuration",
                "username": sms_config.username,
                "sms_config_id": sms_config.id,
                "user_id": sms_config.user.id,
            }, status=status.HTTP_201_CREATED)
        else:
            return Response({
                "status": "failed",
                "errors": [serializer.errors]
            }, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):

        user = request.user

        configurations = SmsConfiguration.objects.filter(user=user)
        serializer = SmsConfigurationSerializerForview(
            configurations, many=True)
        return Response(
            {
                "status": "success",
                "data": serializer.data
            }, status=status.HTTP_200_OK)


class SandboxView(APIView):
    def get(self, request):
        sandbox = SandBox.objects.all()
        serializer = SandboxSerializer(sandbox, many=True).data
        return Response({
            "status": "success",
            "details": serializer
        })


class RecipentsView(APIView):
    def get(self, request):
        recipents = SmsRecipients.objects.all()
        serializer = RecipientsSerializer(recipents, many=True).data
        return Response({
            "status": "success",
            "details": serializer
        })


class SmsComposeView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data
        body = data.get("body")
        config_id = data.get("config_id")
        recipients = data.get("recipients", None)
        user = request.user
        if not all([body, config_id, recipients]):
            return Response(
                {"error": "body, config_id, recipients are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not isinstance(recipients, list) or not recipients:
            return Response({
                "status": "failed",
                "message": "Recipients must be a non-empty list."
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            sms_config = SmsConfiguration.objects.get(id=config_id)
        except (SmsConfiguration.DoesNotExist, User.DoesNotExist):
            return Response({
                "status": "failed",
                "message": "Invalid SMS configuration ID"
            }, status=status.HTTP_404_NOT_FOUND)

        # with transaction.atomic():
        sms_compose = SmsCompose.objects.create(
            body=body,
            sms_configuration=sms_config,
            recipient_number=", ".join(recipients),
            user=user
        )

        for recipient_number in recipients:
            send_sms_task.delay(
                body=body,
                sender_number=sms_config.sender_number,
                recipient_number=recipient_number,
                sms_compose_id=sms_compose.id,
                config_id=sms_config.id

            )

        return Response({
            "status": "processing",
            "sms_compose_id": sms_compose.id,
            "message": "SMS sending tasks have been queued."
        }, status=status.HTTP_202_ACCEPTED)

    def get(self, request):
        user = request.user
        try:
            sms_compose = SmsCompose.objects.filter(user=user)
        except SmsCompose.DoesNotExist:
            return Response({"errors": "Sms Compose user not found"})
        serializer_data = SmsComposeSerializerForView(
            sms_compose, many=True).data
        return Response({
            "status": "success",
            "details": serializer_data
        }, status=status.HTTP_200_OK)




class Recipient_Number_List_View(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            data = request.data

            if 'file' not in data:
                return Response({'status': 'failed', 'errors': 'File not provided.'}, status=400)
            
            file = data['file']

            if not file.name.endswith(('.xlsx', '.xlsm')):
                return Response({'status': 'failed', 'errors': 'Invalid file type.'}, status=400)

            # Read the Excel file
            try:
                wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
            except Exception as e:
                return Response({'status': 'failed', 'errors': f'Error reading Excel file: {str(e)}'}, status=400)

            sheet = wb.active
            Numbers = set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                number = row[0]
                if number:
                    try:
                        number = str(number)
                        Numbers.add(number)
                    except ValueError:
                        continue
            unique_numbers=list(Numbers)
            return Response({'data': unique_numbers})

        except Exception as e:
            return Response({'status': 'failed', 'errors': str(e)}, status=400)
        
class SpecificComposeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, compose_id):
        if not compose_id:
            return Response({"errors": "Compose ID not found"}, status=400)
        try:
            compose = SmsCompose.objects.get(id=compose_id)
        except SmsCompose.DoesNotExist:
            return Response({"errors": "SmsCompose does not exist"}, status=404)

        user = request.user.id
        body = compose.body
        recipient_number_history = []

        recipient_numbers = compose.recipient_number.split(",") 
        recipient_numbers = [number.strip() for number in recipient_numbers]  

        try:
            for number in recipient_numbers:
                history = SmsRecipients.objects.filter(phone_number=number).first()
                if not history:
                    recipient_number_history.append({
                        "number": number,
                        "status": "Not Found",
                        "failed_reason": "Sandbox object not found for this number"
                    })
                else:
                    recipient_number_history.append({
                        "number": number,
                        "status": history.status,
                        "failed_reason": history.failed_reason
                    })

            data = {
                "user_id": user,
                "message_body": body,
                "recipient_history": recipient_number_history
            }
            return Response({"data": data}, status=200)
        except Exception as e:
            return Response({'errors': str(e)}, status=400)
