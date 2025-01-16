from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import load_workbook
from io import BytesIO
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from .models import Email_Address_List, Email_Address
from django.db import transaction
from rest_framework import status
from .serializers import EmailAddressListSerializer


class Email_Address_List_View(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        try:
            user = request.user
            data = request.data
            file = data['file']

            # Read the Excel file
            wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
            sheet = wb.active
            emails = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                email = row[0]
                if email:
                    try:
                        validate_email(email)
                        emails.append(email)
                    except ValidationError as e:
                        print("Error ", e)

            name = user.username
            previous_list = Email_Address_List.objects.filter(
                user=user).order_by("created_at").first()
            if previous_list:
                id = previous_list.id
            else:
                id = 0
            name = f"user.username-list-{id+1}"

            if len(emails) > 0:

                try:
                    email_list = Email_Address_List.objects.create(
                        name=name, user=user)

                    for mail in emails:
                        Email_Address.objects.create(
                            email=mail, email_address_list=email_list)

                    return Response(
                        {
                            'data': emails
                        }
                    )
                except Exception as e:
                    return Response({
                        'status': "failed",
                        'errors': str(e)
                    }, status=status.HTTP_400_BAD_REQUEST)
            return Response({
                'data': []
            })

        except Exception as e:
            return Response({
                'status': "failed",
                'errors': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        user = request.user
        email_address_list = Email_Address_List.objects.filter(user=user)
        serializer = EmailAddressListSerializer(
            email_address_list, many=True)
        return Response({
            'data': serializer.data
        })
